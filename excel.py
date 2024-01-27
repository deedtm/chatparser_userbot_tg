from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import Font, PatternFill

class Excel:
    def __init__(self, file_name: str, sheets_names: list[str], sheets_columns: list[list[str]]):
        self.file_name = file_name + '.xls'
        self.xls = Workbook()
        self.xls.remove(self.xls.active)
        self.add_sheets(sheets_names, sheets_columns)

    def __add_sheet(self, name: str, columns: list[str]):
        sheet: Worksheet = self.xls.create_sheet(name)
        sheet.append(columns)
        self.__styling(sheet, len(columns))

    def add_sheets(self, names: list[str], sheets_columns: list[list[str]]):
        for ind, name in enumerate(names):
            self.__add_sheet(name, sheets_columns[ind])
        self.__save_file()

    def get_sheets_names(self):
        return self.xls.get_sheet_names()
    
    def get_sheet(self, name: str):
        return self.xls.get_sheet_by_name(name)

    def write(self, sheet: Worksheet, datas: list[str]):
        sheet.append(datas)
        self.__save_file()

    def remove_sheet(self, sheet: Worksheet | str):
        xls = self.xls
        if isinstance(Worksheet, sheet):
            xls.remove(sheet)
        elif isinstance(str, sheet):
            sheet = self.get_sheet(sheet)
            xls.remove(sheet)
        else:
            raise TypeError('`sheet` parameter must be `Worksheet` or `str` type')

    def __styling(self, sheet: Worksheet, columns_amount: int):
        font = Font(b=True, color='000000')
        fill = PatternFill('solid', fgColor='FFFF00')

        for letter in 'ABCDEFGHIJKLMN'[:columns_amount]:
            sheet[letter + '1'].font = font
            sheet[letter + '1'].fill = fill

        self.__save_file()

    def __save_file(self):
        self.xls.save(self.file_name)
